import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from camp_headcount_calc import compute

rows, totals = compute()

wb = Workbook()
ws = wb.active
ws.title = '人数確定のご連絡'
ws.sheet_view.rightToLeft = False

FONT_NAME = '游明朝'
THIN = Side(style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', fgColor='D9E2F3')
YELLOW_FILL = PatternFill('solid', fgColor='FFFF00')


def cell(ref, value, bold=False, size=11, align='left', fill=None, border=True, wrap=False):
    c = ws[ref]
    c.value = value
    c.font = Font(name=FONT_NAME, bold=bold, size=size)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fill:
        c.fill = fill
    if border:
        c.border = BORDER
    return c


for col, w in zip('ABCDE', [16, 12, 12, 12, 42]):
    ws.column_dimensions[col].width = w

r = 1
cell(f'A{r}', '夏季合宿　参加人数確定のご連絡', bold=True, size=16, align='center')
ws.merge_cells(f'A{r}:E{r}')
r += 2

cell(f'A{r}', '送信日', bold=True, fill=HEADER_FILL)
cell(f'B{r}', '2026年7月17日', align='left')
ws.merge_cells(f'B{r}:E{r}')
r += 1
cell(f'A{r}', '宛先', bold=True, fill=HEADER_FILL)
cell(f'B{r}', '有限会社　菅平ホテル　御中')
ws.merge_cells(f'B{r}:E{r}')
r += 1
cell(f'A{r}', '発信元', bold=True, fill=HEADER_FILL)
cell(f'B{r}', '駒澤大学高等学校　ラグビー部顧問　占部涼也')
ws.merge_cells(f'B{r}:E{r}')
r += 1
cell(f'A{r}', '連絡先', bold=True, fill=HEADER_FILL)
cell(f'B{r}', 'Tel（携帯）：080-5032-9150／Mail：ryoyaurabe@gmail.com')
ws.merge_cells(f'B{r}:E{r}')
r += 2

cell(f'A{r}', '拝啓　時下ますますご清栄のこととお慶び申し上げます。平素より大変お世話になっております。'
             '2026年6月9日付でいただきましたご請求書（宿泊人数集計表）につきまして、参加生徒の内訳が確定いたしましたので、'
             '下記のとおりご連絡申し上げます。お手数をおかけいたしますが、お食事のご準備の調整にお役立ていただけますと幸いです。',
     wrap=True, border=False)
ws.merge_cells(f'A{r}:E{r}')
ws.row_dimensions[r].height = 60
r += 2

cell(f'A{r}', '１．合宿期間', bold=True, border=False)
r += 1
cell(f'A{r}', '8月1日（土）〜8月7日（金）　6泊7日', border=False)
r += 2

cell(f'A{r}', '２．参加人数の内訳', bold=True, border=False)
r += 1
breakdown = [
    ('部員（生徒）', '42名　（選手39名・マネージャー3名）'),
    ('引率顧問', '2名　（顧問1：全日程帯同／顧問2：8月5日夕食〜合流）'),
    ('外部コーチ', '2名　（入れ替わりで常時1名帯同。1人目：8/1〜8/2、2人目：8/3〜8/5）'),
    ('', '8/6・8/7はコーチの帯同なし'),
    ('最大同時人数', '45名　（部員42名＋顧問2名＋コーチ1名。8/5夕食〜宿泊）'),
]
for label, value in breakdown:
    cell(f'A{r}', label, bold=True, fill=HEADER_FILL if label else None)
    cell(f'B{r}', value)
    ws.merge_cells(f'B{r}:E{r}')
    r += 1
r += 1

cell(f'A{r}', '３．途中合流する部員・顧問について', bold=True, border=False)
r += 1
for line in [
    '椎名　薫（2年F組・マネージャー）：8月4日午後〜合流（夕食からカウント）',
    '吉田　青空（1年E組・選手）：8月5日午後〜合流（夕食からカウント）',
    '顧問2　占部涼也：8月5日午後〜合流（夕食からカウント）',
]:
    cell(f'A{r}', '・' + line, border=False)
    ws.merge_cells(f'A{r}:E{r}')
    r += 1
r += 1

cell(f'A{r}', '４．外部コーチの帯同について', bold=True, border=False)
r += 1
for line in [
    '1人目：8月1日〜8月2日（8/2の夕食・宿泊まで帯同し、翌朝までに交代）',
    '2人目：8月3日〜8月5日（8/5の夕食・宿泊まで帯同）',
    '8月6日・8月7日はコーチの帯同なし',
]:
    cell(f'A{r}', '・' + line, border=False)
    ws.merge_cells(f'A{r}:E{r}')
    r += 1
r += 1

cell(f'A{r}', '５．日別人数表（部員・引率顧問・外部コーチの合計）', bold=True, border=False)
r += 1
headers = ['日付', '朝食', '昼食', '夕食', '宿泊']
for j, h in enumerate(headers):
    cell(f'{get_column_letter(j + 1)}{r}', h, bold=True, align='center', fill=HEADER_FILL)
r += 1
for day, b, l, d, s in rows:
    cell(f'A{r}', day, bold=True, align='center')
    cell(f'B{r}', b, align='center')
    cell(f'C{r}', l, align='center')
    cell(f'D{r}', d, align='center')
    cell(f'E{r}', s, align='center')
    r += 1
cell(f'A{r}', '延べ人数', bold=True, align='center', fill=HEADER_FILL)
cell(f'B{r}', totals['breakfast'], bold=True, align='center', fill=HEADER_FILL)
cell(f'C{r}', totals['lunch'], bold=True, align='center', fill=HEADER_FILL)
cell(f'D{r}', totals['dinner'], bold=True, align='center', fill=HEADER_FILL)
cell(f'E{r}', totals['stay'], bold=True, align='center', fill=HEADER_FILL)
r += 2

cell(f'A{r}', '※　上表は部員42名（途中合流2名を含む）・引率顧問2名（顧問2は8/5夕食〜）・'
             '外部コーチ（8/1〜8/5、入れ替わりで常時1名）を合算した確定人数です。'
             'ご請求書（6/9付）でいただいた46名／日の一律見積と比べ、日によって人数が変動いたしますので、'
             '上表の人数でご準備をお願いいたします。',
     size=9.5, wrap=True, border=False)
ws.merge_cells(f'A{r}:E{r}')
ws.row_dimensions[r].height = 45
r += 2

cell(f'A{r}', '６．お願い', bold=True, border=False)
r += 1
cell(f'A{r}', '上記人数にもとづき、お食事の提供数調整をお願いできますと幸いです。'
             'なお、最終のご請求額につきましては、合宿終了後の実績人数にもとづき、'
             '精算いただけますようお願い申し上げます。ご不明な点がございましたら、上記連絡先までご連絡ください。',
     wrap=True, border=False)
ws.merge_cells(f'A{r}:E{r}')
ws.row_dimensions[r].height = 45
r += 2

cell(f'A{r}', '敬具', border=False, align='right')
ws.merge_cells(f'A{r}:E{r}')

out_path = '/home/user/teacher-automation-lab/output/2026_夏合宿_人数確定のご連絡_菅平ホテル様.xlsx'
os.makedirs('/home/user/teacher-automation-lab/output', exist_ok=True)
wb.save(out_path)
print(f'保存完了: {out_path}')
print(totals)
